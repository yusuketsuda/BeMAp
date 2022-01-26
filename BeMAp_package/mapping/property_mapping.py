#!/usr/bin/python

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

def num_to_alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num_to_alpha(num//26-1)+chr(90)
    else:
        return num_to_alpha(num//26)+chr(64+num%26)

def make_mapping(new_AMR, new_centered_gene, color, prop_color, summary, prop_name, legend1, legend2, dendrogram, out=False):
    
    prop = pd.DataFrame(summary.loc[:, prop_name], index=list(summary.index))

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']

    sheet.freeze_panes = 'A6'

    # make white line
    side = Side(style='thin', color='ffffff')
    border = Border(top=side, bottom=side,right=side,left=side)
            
    # input accession numbers and genes
    for i in range(len(new_AMR.columns)):
        sheet.cell(row = 4, column = i+1, value=new_centered_gene.columns[i])
        sheet.cell(row = 4, column = i+1).alignment = Alignment(horizontal = 'center', vertical = 'top',textRotation=180)
        sheet.cell(row = 4, column = i+1).border = border
        for j in range(len(new_AMR.index)):
            sheet.cell(row = j+6, column = i+1, value=new_centered_gene.iat[j,i])
            sheet.cell(row = j+6, column = i+1).fill = openpyxl.styles.PatternFill(patternType = 'solid', fgColor ='f7f7f7', bgColor = 'f7f7f7')
            sheet.cell(row = j+6, column = i+1).font = Font(size = 1)
            sheet.cell(row = j+6, column = i+1).border = border
            sheet.column_dimensions[num_to_alpha(j+1)].width = 2
            
    # make squared cells
    for i in range(sheet.max_row):
        sheet.row_dimensions[i+6].height = 11

    # make color spreadsheet                                                                            
    all_gray = np.full(len(new_AMR.index)*len(new_AMR.columns),'f7f7f7').reshape(len(new_AMR.index),len(new_AMR.columns))
    spreadsheet_color = pd.DataFrame(all_gray,  index=new_AMR.index, columns=new_AMR.columns)

    # coloring according to each propety of datasets
    for h in range(len(spreadsheet_color.columns)):
        sheet.cell(row= 5, column = h+1, value = prop.loc[new_AMR.columns[h],prop_name])
        sheet.cell(row = 5, column = h+1).alignment = Alignment(horizontal = 'center', vertical = 'top',textRotation=180)
        sheet.cell(row = 5, column = h+1).border = border

        if prop_name == 'country' or prop_name == 'organism':
            for i in prop_color.index:
                if isinstance(prop.loc[new_AMR.columns[h],prop_name], str):
                    if prop_color.iloc[i,0] in prop.loc[new_AMR.columns[h],prop_name]:
                        spreadsheet_color.iloc[:,h] = prop_color.iloc[i,1]
                        sheet.cell(row= 5, column=h+1).fill = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = prop_color.iloc[i,1], bgColor = prop_color.iloc[i,1])
        else:
            for i in prop_color.index:
                if prop.loc[new_AMR.columns[h],prop_name] == prop_color.iloc[i,0]:
                    sheet.cell(row= 5, column=h+1).fill = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = prop_color.iloc[i,1], bgColor = prop_color.iloc[i,1])
                    spreadsheet_color.iloc[:,h] = prop_color.iloc[i,1]

    for h in color.index:
        spreadsheet_color[new_AMR == h] = color.loc[h,1]
    
    spreadsheet_color[new_AMR.isnull()] = 'ffffff'
    
    for i in range(len(spreadsheet_color.index)):
        for j in range(len(spreadsheet_color.columns)):
            sheet.cell(row=i+6, column=j+1).fill = openpyxl.styles.PatternFill(patternType = 'solid',
                                                                               fgColor = spreadsheet_color.iat[i,j],
                                                                               bgColor = spreadsheet_color.iat[i,j])
    
    # show legend in A1 
    sheet.row_dimensions[1].height = 30
    legend = openpyxl.drawing.image.Image(legend1)
    w = legend.width
    h = legend.height
    legend.width = w*40.965/h
    legend.height = 40.965
    sheet.add_image(legend,'A1')

    # show legend for property in A2
    sheet.row_dimensions[2].height = 30
    legend = openpyxl.drawing.image.Image(legend2)
    w = legend.width
    h = legend.height
    legend.width = w*40.965/h
    legend.height = 40.965
    sheet.add_image(legend,'A2')

    # show dendrogram in A3
    sheet.row_dimensions[3].height = 50   
    img = openpyxl.drawing.image.Image(dendrogram)
    img.width = (len(new_AMR.columns) * 0.1663-0.005)/0.0104            # width of each cell 0.1663、function for showing image Y=0.0104x + 0.005, the number of cells
    img.height = 68.26
    sheet.add_image(img, 'A3')
    
    if out:
        wb.save(out + '/' + prop_name + '.xlsx')
    else:
        wb.save(prop_name + '.xlsx')
